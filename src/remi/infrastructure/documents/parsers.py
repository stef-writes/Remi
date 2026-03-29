"""Document parsers — CSV and Excel file ingestion."""

from __future__ import annotations

import csv
import io
import re
import uuid
from typing import Any

from remi.domain.documents.models import Document

# Patterns that indicate a row is report metadata (title, export info, filters)
# rather than a real column header row.
_METADATA_PATTERNS = re.compile(
    r"^(exported\s+on|properties:|units:|date\s+range|filter|exclude|include|"
    r"base\s+report|property\s+groups|as\s+of|level\s+of\s+detail|bedrooms|"
    r"bathrooms|amenities|appliances|report\s+builder|balance:|amount\s+owed)",
    re.I,
)

# Minimum number of non-null cells for a row to be considered a real header.
_MIN_HEADER_CELLS = 3


def parse_csv(filename: str, content: bytes | str) -> Document:
    """Parse a CSV file into a Document with typed rows."""
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")

    reader = csv.DictReader(io.StringIO(content))
    columns = reader.fieldnames or []
    rows: list[dict[str, Any]] = []
    for row in reader:
        cleaned = {k: _coerce_value(v) for k, v in row.items() if k is not None}
        rows.append(cleaned)

    return Document(
        id=f"doc-{uuid.uuid4().hex[:12]}",
        filename=filename,
        content_type="text/csv",
        row_count=len(rows),
        column_names=list(columns),
        rows=rows,
    )


def parse_excel(filename: str, content: bytes) -> Document:
    """Parse an Excel file (.xlsx) into a Document.

    Handles report-style exports (AppFolio, Yardi, etc.) that prepend several
    rows of metadata before the real column header.  The parser scans forward
    to find the first row that looks like a genuine header, skipping title /
    filter rows automatically.

    Requires the openpyxl optional dependency.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel parsing. Install with: pip install remi[documents]"
        ) from exc

    _CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return Document(
            id=f"doc-{uuid.uuid4().hex[:12]}",
            filename=filename,
            content_type=_CONTENT_TYPE,
        )

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return Document(
            id=f"doc-{uuid.uuid4().hex[:12]}",
            filename=filename,
            content_type=_CONTENT_TYPE,
        )

    header_idx = _find_header_row(all_rows)
    header = all_rows[header_idx]
    columns = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]

    rows: list[dict[str, Any]] = []
    for raw_row in all_rows[header_idx + 1 :]:
        # Skip completely empty rows (section dividers / totals spacers)
        if all(c is None for c in raw_row):
            continue
        row_dict = {}
        for col_name, val in zip(columns, raw_row):
            row_dict[col_name] = _coerce_value(val)
        rows.append(row_dict)

    return Document(
        id=f"doc-{uuid.uuid4().hex[:12]}",
        filename=filename,
        content_type=_CONTENT_TYPE,
        row_count=len(rows),
        column_names=columns,
        rows=rows,
    )


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    """Return the index of the first row that looks like a real column header.

    Strategy:
    1. Skip rows that are clearly report metadata (match _METADATA_PATTERNS or
       have only one non-null cell).
    2. Accept the first row with >= _MIN_HEADER_CELLS non-null string cells
       that does NOT look like a data row (i.e. no datetime / numeric values
       in the first few cells — those belong to data rows).
    """
    for idx, row in enumerate(rows):
        non_null = [c for c in row if c is not None]
        if len(non_null) < _MIN_HEADER_CELLS:
            continue

        first_val = str(non_null[0]).strip() if non_null else ""

        # Skip known metadata lines
        if _METADATA_PATTERNS.match(first_val):
            continue

        # Skip rows where the first cell is a datetime (data row, not header)
        import datetime as _dt
        if isinstance(non_null[0], (_dt.datetime, _dt.date)):
            continue

        # Skip rows where the majority of non-null cells are numeric (data rows)
        numeric_count = sum(1 for c in non_null if isinstance(c, (int, float)))
        if numeric_count > len(non_null) // 2:
            continue

        # This row looks like a header
        return idx

    # Fallback: use row 0
    return 0


def _coerce_value(val: Any) -> Any:
    """Best-effort coercion: numbers, booleans, nulls."""
    if val is None:
        return None
    if not isinstance(val, str):
        return val
    stripped = val.strip()
    if stripped == "":
        return None
    if stripped.lower() in ("true", "yes"):
        return True
    if stripped.lower() in ("false", "no"):
        return False
    try:
        return int(stripped)
    except ValueError:
        pass
    try:
        return float(stripped)
    except ValueError:
        pass
    return stripped
